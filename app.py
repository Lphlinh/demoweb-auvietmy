import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import calendar
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import io
import openpyxl
from openpyxl.styles import Font

st.set_page_config(page_title="Hệ thống Quản lý Âu Việt Mỹ", layout="wide", page_icon="🛡️")

# ==========================================
# 1. KẾT NỐI GOOGLE SHEETS
# ==========================================
@st.cache_resource
def init_connection():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    try:
        creds_dict = json.loads(st.secrets["GOOGLE_CREDENTIALS"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    client = gspread.authorize(creds)
    return client.open_by_key("13Y44fuaCvd1yTZvlMzTtFoyFpfOLb-PoLTrcvkEEICY")

sheet = init_connection()

# ==========================================
# 2. HÀM QUÉT TKB & XỬ LÝ LỊCH
# ==========================================
@st.cache_data(ttl=600) 
def load_master_data():
    ds_gv = pd.DataFrame(sheet.worksheet("DS_GV").get_all_records())
    try:
        tkb_raw = sheet.worksheet("TKB_PhanCong").get_all_values()
        df_tkb = pd.DataFrame(tkb_raw)
        classes = ["KHTN", "KHXH", "11A", "11C", "10A", "10C", "9A", "8A", "7A", "6A"]
        pc_data = []
        current_thu, current_tiet = "", ""
        for row_idx in range(7, df_tkb.shape[0]): 
            val_thu = str(df_tkb.iloc[row_idx, 0]).strip()
            if val_thu:
                if val_thu == "2" or "thứ 2" in val_thu.lower() or "thứ hai" in val_thu.lower(): current_thu = "Thứ Hai"
                elif val_thu == "3" or "thứ 3" in val_thu.lower() or "thứ ba" in val_thu.lower(): current_thu = "Thứ Ba"
                elif val_thu == "4" or "thứ 4" in val_thu.lower() or "thứ tư" in val_thu.lower(): current_thu = "Thứ Tư"
                elif val_thu == "5" or "thứ 5" in val_thu.lower() or "thứ năm" in val_thu.lower(): current_thu = "Thứ Năm"
                elif val_thu == "6" or "thứ 6" in val_thu.lower() or "thứ sáu" in val_thu.lower(): current_thu = "Thứ Sáu"
                elif val_thu == "7" or "thứ 7" in val_thu.lower() or "thứ bảy" in val_thu.lower(): current_thu = "Thứ Bảy"
            val_tiet = str(df_tkb.iloc[row_idx, 1]).strip()
            if val_tiet: current_tiet = val_tiet
            for col_idx, class_name in enumerate(classes, start=2): 
                cell = df_tkb.iloc[row_idx, col_idx]
                if cell and "-" in cell:
                    try:
                        mon, gv_short = cell.split("-")
                        name_clean = gv_short.replace("T.", "").replace("C.", "").strip()
                        match = ds_gv[ds_gv['Họ tên Giáo viên'].str.contains(name_clean, case=False, na=False)]
                        if not match.empty:
                            pc_data.append({
                                "Lớp": class_name, "Môn học": mon.strip(),
                                "Họ tên GV": match.iloc[0]['Họ tên Giáo viên'],
                                "Mã định danh": str(match.iloc[0]['Mã định danh']),
                                "Thứ": current_thu, "Tiết": current_tiet
                            })
                    except: pass
        pc_chuyenmon = pd.DataFrame(pc_data).drop_duplicates()
    except:
        pc_chuyenmon = pd.DataFrame(columns=["Lớp", "Môn học", "Họ tên GV", "Mã định danh", "Thứ", "Tiết"])
    return ds_gv, pc_chuyenmon

ds_gv, pc_chuyenmon = load_master_data()

def get_month_calendar(year, month):
    cal = calendar.monthcalendar(year, month)
    weeks = []
    for week in cal:
        days = [d for d in week if d != 0]
        if days:
            start_date = f"{days[0]:02d}/{month:02d}"
            end_date = f"{days[-1]:02d}/{month:02d}"
            weeks.append({"days": week, "title": f"{start_date} - {end_date}"})
    return weeks

# ==========================================
# 3. MÀN HÌNH ĐĂNG NHẬP
# ==========================================
if "logged_in" not in st.session_state:
    st.session_state.update({"logged_in": False, "role": None, "user_name": None, "user_id": None})

if not st.session_state.logged_in:
    st.markdown("<h1 style='text-align: center;'>🛡️ CỔNG ĐĂNG NHẬP ÂU VIỆT MỸ</h1>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        loai_tk = st.selectbox("Vai trò của bạn:", ["Giáo viên", "Giám thị", "Ban Giám Hiệu"])
        mat_khau = st.text_input("Mật khẩu / Mã định danh:", type="password")
        if st.button("Đăng nhập", use_container_width=True):
            if loai_tk == "Giám thị":
                try: pass_gt = st.secrets["PASS_GT"]
                except: pass_gt = "giamthi123"
                if mat_khau == pass_gt:
                    st.session_state.update({"logged_in": True, "role": "Giám thị", "user_name": "Tổ Giám thị"})
                    st.rerun()
                else: st.error("Sai mật khẩu!")
            elif loai_tk == "Ban Giám Hiệu":
                try: pass_bgh = st.secrets["PASS_BGH"]
                except: pass_bgh = "hieutruong123"
                if mat_khau == pass_bgh:
                    st.session_state.update({"logged_in": True, "role": "BGH", "user_name": "Ban Giám Hiệu"})
                    st.rerun()
                else: st.error("Sai mật khẩu!")
            elif loai_tk == "Giáo viên":
                gv_match = ds_gv[ds_gv['Mã định danh'].astype(str) == mat_khau.strip()]
                if not gv_match.empty:
                    st.session_state.update({"logged_in": True, "role": "Giáo viên", 
                                             "user_name": gv_match.iloc[0]['Họ tên Giáo viên'], "user_id": mat_khau.strip()})
                    st.rerun()
                else: st.error("Mã định danh không tồn tại!")
else:
    with st.sidebar:
        st.success(f"👤 **{st.session_state.user_name}**")
        if st.button("🚪 Đăng xuất", use_container_width=True):
            st.session_state.logged_in = False
            st.rerun()

    # ==========================================
    # 4. CHỨC NĂNG GIÁM THỊ (CÓ BÁO CÁO CUỐI TUẦN)
    # ==========================================
    if st.session_state.role == "Giám thị":
        tab_gt1, tab_gt2 = st.tabs(["📝 Ghi nhận biến động", "🔎 Báo cáo Cuối tuần"])
        
        with tab_gt1:
            st.header("Ghi nhận sự cố")
            col_date, _ = st.columns([1, 2])
            with col_date:
                ngay_chon = st.date_input("🗓️ Chọn ngày ghi nhận:", value=datetime.now().date())
                ngay_str = ngay_chon.strftime("%d/%m/%Y")
                thu_hien_tai = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"][ngay_chon.weekday()]
            
            if ngay_chon.weekday() == 6:
                st.error("🔒 HỆ THỐNG ĐÃ KHÓA SỔ. Hôm nay là Chủ Nhật, không thể cập nhật dữ liệu.")
            else:
                df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                df_today = df_ngoai_le[df_ngoai_le['Ngày'] == ngay_str] if not df_ngoai_le.empty else pd.DataFrame()
                tkb_today = pc_chuyenmon[pc_chuyenmon['Thứ'] == thu_hien_tai]

                if pc_chuyenmon.empty:
                    st.error("❌ App chưa nhận diện được dữ liệu từ TKB.")
                else:
                    st.markdown("---")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        classes = ["KHTN", "KHXH", "11A", "11C", "10A", "10C", "9A", "8A", "7A", "6A"]
                        lop = st.selectbox("Lớp", classes)
                        mon_hople = tkb_today[tkb_today['Lớp'] == lop]['Môn học'].dropna().unique().tolist()
                    if not mon_hople:
                        st.warning(f"📭 Vào {thu_hien_tai} ({ngay_str}), lớp {lop} KHÔNG CÓ lịch.")
                    else:
                        with col1:
                            mon = st.selectbox("Môn", mon_hople)
                            tiet_hop_le = sorted([str(t) for t in tkb_today[(tkb_today['Lớp'] == lop) & (tkb_today['Môn học'] == mon)]['Tiết'].dropna().unique().tolist()])
                            tiet_list = st.multiselect("Chọn Tiết", options=tiet_hop_le, default=tiet_hop_le)

                        gv_info = tkb_today[(tkb_today['Lớp'] == lop) & (tkb_today['Môn học'] == mon)]
                        gv_goc_ten = gv_info.iloc[0]['Họ tên GV'] if not gv_info.empty else "N/A"
                        gv_goc_id = str(gv_info.iloc[0]['Mã định danh']) if not gv_info.empty else ""
                        with col2:
                            st.info(f"GV Phụ trách: **{gv_goc_ten}**")
                            loai = st.selectbox("Loại", ["Nghỉ có phép", "Nghỉ không phép", "Dạy thay", "Đổi tiết"])
                        
                        gv_ban_list = []
                        for t in tiet_list:
                            gv_ban_list.extend(tkb_today[tkb_today['Tiết'] == str(t)]['Mã định danh'].astype(str).tolist())
                            if not df_today.empty:
                                ca_nay = df_today[df_today['Tiết'].astype(str) == str(t)]
                                gv_ban_list.extend(ca_nay['ID GV vắng'].astype(str).tolist())
                                gv_ban_list.extend(ca_nay['ID GV dạy thay'].astype(str).tolist())
                        
                        if gv_goc_id: gv_ban_list.append(gv_goc_id) 
                        gv_ban_list = list(set([x for x in gv_ban_list if x != ""]))
                        df_gv_ranh = ds_gv[~ds_gv['Mã định danh'].astype(str).isin(gv_ban_list)]
                        danh_sach_day_thay = ["Không"] + df_gv_ranh['Họ tên Giáo viên'].tolist()

                        with col3:
                            gv_thay_ten = st.selectbox("GV Dạy thay", danh_sach_day_thay)
                            gv_thay_id = str(ds_gv[ds_gv['Họ tên Giáo viên'] == gv_thay_ten]['Mã định danh'].values[0]) if gv_thay_ten != "Không" else ""
                            note = st.text_area("Ghi chú")

                        if st.button("💾 Lưu báo cáo", type="primary"):
                            if len(tiet_list) == 0:
                                st.warning("⚠️ Vui lòng chọn ít nhất 1 tiết học!")
                            else:
                                with st.spinner("Đang lưu dữ liệu..."):
                                    rows_to_add = [[ngay_str, thu_hien_tai, t, lop, mon, loai, gv_goc_id, gv_thay_id, note] for t in tiet_list]
                                    sheet.worksheet("BaoCao_NgoaiLe").append_rows(rows_to_add)
                                    st.success(f"✅ Đã ghi nhận thành công cho ngày {ngay_str}!")

        with tab_gt2:
            st.subheader("Báo cáo Kiểm dò chéo Sổ đầu bài")
            st.info("💡 Tính năng này giúp Giám thị rà soát lại số liệu thực tế so với TKB gốc.")
            col_d1, col_d2 = st.columns(2)
            with col_d1: start_rp = st.date_input("Từ ngày:", value=datetime.now().date() - timedelta(days=datetime.now().weekday()))
            with col_d2: end_rp = st.date_input("Đến ngày:", value=start_rp + timedelta(days=6))
            
            if st.button("Tạo Báo cáo Tuần"):
                df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                if df_ngoai_le.empty: st.warning("Chưa có dữ liệu ngoại lệ.")
                else:
                    df_ngoai_le['Ngày chuẩn'] = pd.to_datetime(df_ngoai_le['Ngày'], format='%d/%m/%Y', errors='coerce')
                    mask_rp = (df_ngoai_le['Ngày chuẩn'].dt.date >= start_rp) & (df_ngoai_le['Ngày chuẩn'].dt.date <= end_rp)
                    df_rp = df_ngoai_le.loc[mask_rp].copy()
                    
                    tkb_tuan = pc_chuyenmon.copy()
                    tkb_tuan['Khối'] = "Lớp " + tkb_tuan['Lớp'].astype(str).str.extract(r'^(\d+)')[0].fillna("Khác")
                    rp_tkb = tkb_tuan.groupby('Khối').size().reset_index(name='Tổng TKB phải dạy')
                    
                    if not df_rp.empty:
                        df_rp['Khối'] = "Lớp " + df_rp['Lớp'].astype(str).str.extract(r'^(\d+)')[0].fillna("Khác")
                        rp_vang = df_rp[df_rp['ID GV vắng'] != ""].groupby('Khối').size().reset_index(name='Số tiết Nghỉ (Vắng)')
                        rp_thay = df_rp[df_rp['ID GV dạy thay'] != ""].groupby('Khối').size().reset_index(name='Số tiết Dạy thay')
                    else:
                        rp_vang = pd.DataFrame(columns=['Khối', 'Số tiết Nghỉ (Vắng)'])
                        rp_thay = pd.DataFrame(columns=['Khối', 'Số tiết Dạy thay'])
                    
                    rp_final = pd.merge(rp_tkb, rp_vang, on='Khối', how='left').fillna(0)
                    rp_final = pd.merge(rp_final, rp_thay, on='Khối', how='left').fillna(0)
                    rp_final['Tổng Thực Dạy'] = rp_final['Tổng TKB phải dạy'] - rp_final['Số tiết Nghỉ (Vắng)'] + rp_final['Số tiết Dạy thay']
                    
                    rp_final.loc['TOÀN TRƯỜNG'] = rp_final.sum(numeric_only=True)
                    rp_final.at['TOÀN TRƯỜNG', 'Khối'] = "TOÀN TRƯỜNG"
                    st.dataframe(rp_final, use_container_width=True)

    # ==========================================
    # 5. CHỨC NĂNG GIÁO VIÊN
    # ==========================================
    elif st.session_state.role == "Giáo viên":
        st.header(f"🔍 Hồ sơ đối soát của Thầy/Cô: {st.session_state.user_name}")
        df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
        if not df_ngoai_le.empty:
            gv_id_str = str(st.session_state.user_id)
            df_vang = df_ngoai_le[df_ngoai_le['ID GV vắng'].astype(str) == gv_id_str].copy()
            if not df_vang.empty: df_vang['Vai trò'] = "Vắng mặt (-)"
            df_thay = df_ngoai_le[df_ngoai_le['ID GV dạy thay'].astype(str) == gv_id_str].copy()
            if not df_thay.empty: df_thay['Vai trò'] = "Dạy thay (+)"
            df_ketqua = pd.concat([df_vang, df_thay])
            if not df_ketqua.empty:
                st.dataframe(df_ketqua[['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Vai trò', 'Loại ngoại lệ']], use_container_width=True)
            else: st.info("🎉 Thầy/Cô đảm bảo 100% công giảng dạy.")
        else: st.info("Hệ thống hiện chưa có dữ liệu.")

    # ==========================================
    # 6. CHỨC NĂNG BGH (DASHBOARD TỔNG QUÁT & EXCEL)
    # ==========================================
    elif st.session_state.role == "BGH":
        st.header("📊 Bảng điều khiển dành cho Ban Giám Hiệu")
        
        df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
        if df_ngoai_le.empty:
            min_date, max_date = datetime.now().date(), datetime.now().date()
            df_filtered = pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])
        else:
            df_ngoai_le['Ngày chuẩn'] = pd.to_datetime(df_ngoai_le['Ngày'], format='%d/%m/%Y', errors='coerce')
            min_date = df_ngoai_le['Ngày chuẩn'].min().date()
            max_date = df_ngoai_le['Ngày chuẩn'].max().date()
            date_range = st.date_input("🗓️ Chọn khoảng thời gian xem báo cáo:", value=(min_date, max_date))
            
            if isinstance(date_range, tuple) and len(date_range) == 2: start_date, end_date = date_range
            elif isinstance(date_range, tuple) and len(date_range) == 1: start_date = end_date = date_range[0]
            else: start_date = end_date = date_range
                
            mask = (df_ngoai_le['Ngày chuẩn'].dt.date >= start_date) & (df_ngoai_le['Ngày chuẩn'].dt.date <= end_date)
            df_filtered = df_ngoai_le.loc[mask].copy()

        dict_gv_ten = pd.Series(ds_gv['Họ tên Giáo viên'].values, index=ds_gv['Mã định danh'].astype(str)).to_dict()
        dict_gv_to = pd.Series(ds_gv['Tổ chuyên môn'].values, index=ds_gv['Mã định danh'].astype(str)).to_dict()
        
        if not df_filtered.empty:
            df_filtered['Giáo viên Vắng'] = df_filtered['ID GV vắng'].astype(str).map(dict_gv_ten).fillna("Không rõ")
            df_filtered['Tổ Vắng'] = df_filtered['ID GV vắng'].astype(str).map(dict_gv_to).fillna("Không rõ")
            df_filtered['Giáo viên Dạy thay'] = df_filtered['ID GV dạy thay'].astype(str).map(dict_gv_ten).fillna("Không có")
            df_filtered['Tổ Dạy thay'] = df_filtered['ID GV dạy thay'].astype(str).map(dict_gv_to).fillna("Không có")

        tab1, tab2, tab3, tab4 = st.tabs(["📊 Tổng quát", "🏢 Theo Tổ", "📥 Xuất EXCEL (Chấm Công)", "🗂️ TKB Đã Quét"])
        
        with tab1:
            st.subheader("Nhật ký Biến động Tổng quát")
            tong_su_co = len(df_filtered)
            so_ca_day_thay = len(df_filtered[df_filtered['ID GV dạy thay'] != '']) if not df_filtered.empty else 0
            col1, col2, col3 = st.columns(3)
            col1.metric("Tổng tiết báo vắng", tong_su_co)
            col2.metric("Số tiết đã Dạy thay", so_ca_day_thay, delta_color="normal")
            col3.metric("Số tiết Lớp tự học", tong_su_co - so_ca_day_thay, delta_color="inverse")
            if not df_filtered.empty:
                st.dataframe(df_filtered[['Ngày', 'Tiết', 'Lớp', 'Môn', 'Giáo viên Vắng', 'Giáo viên Dạy thay', 'Ghi chú']], use_container_width=True)
            else: st.info("Trường đang hoạt động ổn định, chưa phát sinh ca vắng/dạy thay nào.")
        
        with tab2:
            st.subheader("Thống kê theo Tổ Chuyên môn")
            if not df_filtered.empty:
                df_to_vang = df_filtered.groupby('Tổ Vắng').size().reset_index(name='Số tiết Vắng')
                df_to_thay = df_filtered[df_filtered['Tổ Dạy thay'] != "Không có"].groupby('Tổ Dạy thay').size().reset_index(name='Số tiết Dạy thay')
                df_to_tonghop = pd.merge(df_to_vang, df_to_thay, left_on='Tổ Vắng', right_on='Tổ Dạy thay', how='outer').fillna(0)
                df_to_tonghop['Tổ chuyên môn'] = df_to_tonghop['Tổ Vắng'].combine_first(df_to_tonghop['Tổ Dạy thay'])
                st.dataframe(df_to_tonghop[['Tổ chuyên môn', 'Số tiết Vắng', 'Số tiết Dạy thay']], use_container_width=True)
            else: st.info("Không có dữ liệu để thống kê theo tổ.")

        with tab4:
            st.subheader("Dữ liệu Thời Khóa Biểu đang áp dụng")
            st.dataframe(pc_chuyenmon, use_container_width=True)

        with tab3:
            st.subheader("Tạo Bảng Chấm Công Lương (Mẫu Kế toán)")
            col_m, col_y = st.columns(2)
            with col_m: thang_xuat = st.selectbox("Chọn Tháng:", range(1, 13), index=datetime.now().month - 1)
            with col_y: nam_xuat = st.selectbox("Chọn Năm:", [2024, 2025, 2026, 2027], index=2)
            
            weeks = get_month_calendar(nam_xuat, thang_xuat)
            ds_gv['HienThi_BGH'] = ds_gv['Họ tên Giáo viên'] + " - ID: " + ds_gv['Mã định danh'].astype(str)
            gv_chon = st.selectbox("Chọn Giáo viên để xuất Excel:", ["-- Chọn Giáo viên --"] + ds_gv['HienThi_BGH'].tolist())
            
            def tao_excel_mau_avm(gv_dict, weeks, month, year, df_tkb_all, df_nl_all):
                try:
                    wb = openpyxl.load_workbook("BaoCaoMau.xlsx")
                except FileNotFoundError:
                    st.error("❌ Không tìm thấy file 'BaoCaoMau.xlsx'. Vui lòng upload template này lên cùng thư mục với app.py.")
                    return None

                template_ws = wb.active
                template_ws_name = template_ws.title

                danh_sach_thu = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy"]
                
                font_v = Font(color="FF0000", bold=True)
                font_dt = Font(color="00B050", bold=True)

                for gv_id, gv_name in gv_dict.items():
                    tkb_gv = df_tkb_all[df_tkb_all['Mã định danh'].astype(str) == gv_id]
                    nl_gv_v = df_nl_all[df_nl_all['ID GV vắng'].astype(str) == gv_id]
                    nl_gv_dt = df_nl_all[df_nl_all['ID GV dạy thay'].astype(str) == gv_id]

                    if tkb_gv.empty and nl_gv_v.empty and nl_gv_dt.empty: continue 

                    ws = wb.copy_worksheet(template_ws)
                    ws.title = gv_name[:31]

                    ws['H1'] = f'Giáo viên: {gv_name}'
                    ws['A3'] = f'THÁNG {month} NĂM {year}'

                    for w_idx, w in enumerate(weeks):
                        ws.cell(row=4, column=5 + (w_idx * 3), value=w['title'])

                    for thu in danh_sach_thu:
                        thu_idx = danh_sach_thu.index(thu)
                        tkb_thu = tkb_gv[tkb_gv['Thứ'] == thu]

                        for tiet in range(1, 9): 
                            # TÍNH TOÁN DÒNG CHUẨN XÁC:
                            # Base của các ngày: 7, 17, 27, 37, 47, 57 (Cách nhau 10 dòng)
                            # Bỏ qua 1 dòng nghỉ giữa buổi nếu từ tiết 5 trở đi
                            base_row_ngay = 7 + (thu_idx * 10)
                            offset_tiet = (tiet - 1) if tiet <= 4 else tiet
                            row_idx = base_row_ngay + offset_tiet
                            
                            base_class = ""
                            tkb_match = tkb_thu[tkb_thu['Tiết'] == str(tiet)]
                            if not tkb_match.empty: base_class = tkb_match.iloc[0]['Lớp']

                            for w_idx, w in enumerate(weeks):
                                col_idx = 5 + (w_idx * 3) 
                                day = w['days'][thu_idx]

                                if day != 0:
                                    ngay_str = f"{day:02d}/{month:02d}/{year}"
                                    nl_v_match = nl_gv_v[(nl_gv_v['Ngày'] == ngay_str) & (nl_gv_v['Tiết'].astype(str) == str(tiet))]
                                    nl_dt_match = nl_gv_dt[(nl_gv_dt['Ngày'] == ngay_str) & (nl_gv_dt['Tiết'].astype(str) == str(tiet))]
                                    
                                    target_cell = ws.cell(row=row_idx, column=col_idx)
                                    
                                    if not nl_v_match.empty:
                                        target_cell.value = f"V ({base_class})"
                                        target_cell.font = font_v
                                    elif not nl_dt_match.empty:
                                        lop_dt = nl_dt_match.iloc[0]['Lớp']
                                        target_cell.value = f"{lop_dt} (DT)" 
                                        target_cell.font = font_dt
                                    else:
                                        target_cell.value = base_class
                                else:
                                    ws.cell(row=row_idx, column=col_idx, value="-")

                wb.remove(wb[template_ws_name])

                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()

            st.markdown("---")
            col_ex1, col_ex2 = st.columns(2)
            df_nl_full = df_ngoai_le.copy() if not df_ngoai_le.empty else pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])
            
            with col_ex1:
                if gv_chon != "-- Chọn Giáo viên --":
                    if st.button(f"📥 Tải Excel CÁ NHÂN ({gv_chon.split(' - ')[0]})", type="primary"):
                        with st.spinner("Đang tạo Excel từ Template..."):
                            gv_id_str = gv_chon.split(" - ID: ")[-1].strip()
                            gv_name_str = gv_chon.split(" - ID: ")[0].strip()
                            excel_data = tao_excel_mau_avm({gv_id_str: gv_name_str}, weeks, thang_xuat, nam_xuat, pc_chuyenmon, df_nl_full)
                            if excel_data:
                                st.download_button(label="✅ Bấm để tải File", data=excel_data, file_name=f"ChamCong_{gv_name_str}_{thang_xuat}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            with col_ex2:
                if st.button("📥 Tải Excel TOÀN TRƯỜNG (Tất cả GV)", type="primary"):
                    with st.spinner("Đang tổng hợp dữ liệu toàn trường..."):
                        gv_dict_all = pd.Series(ds_gv['Họ tên Giáo viên'].values, index=ds_gv['Mã định danh'].astype(str)).to_dict()
                        excel_data_all = tao_excel_mau_avm(gv_dict_all, weeks, thang_xuat, nam_xuat, pc_chuyenmon, df_nl_full)
                        if excel_data_all:
                            st.download_button(label="✅ Đã xong! Tải File", data=excel_data_all, file_name=f"ChamCong_ToanTruong_T{thang_xuat}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")